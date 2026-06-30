import csv
import os
import re
import tempfile
from pathlib import Path
from typing import Any

from docx import Document as DocxDocument
from fastapi import FastAPI
from minio import Minio
from neo4j import GraphDatabase
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from pypdf import PdfReader

app = FastAPI(title="Nornickel GraphRAG Service")

NEO4J_URI = os.getenv("NEO4J_URI", "bolt://localhost:7687")
NEO4J_USERNAME = os.getenv("NEO4J_USERNAME", "neo4j")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD", "nornickel-password")
MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT", "localhost:9000")
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY", "minioadmin")
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY", "minioadmin")
MINIO_BUCKET = os.getenv("MINIO_BUCKET", "documents")

driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USERNAME, NEO4J_PASSWORD))
minio_client = Minio(
    MINIO_ENDPOINT.replace("http://", "").replace("https://", ""),
    access_key=MINIO_ACCESS_KEY,
    secret_key=MINIO_SECRET_KEY,
    secure=MINIO_ENDPOINT.startswith("https://"),
)


class EntityMention(BaseModel):
    id: str
    type: str
    label: str


class RetrieveRequest(BaseModel):
    query: str
    mentions: list[EntityMention] = Field(default_factory=list)


class ExtractRequest(BaseModel):
    documentId: str
    title: str
    type: str
    storageKey: str | None = None


class PublishRequest(BaseModel):
    extraction: dict[str, Any]


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/internal/graphrag/retrieve")
def retrieve(request: RetrieveRequest) -> dict[str, Any]:
    rows = query_graph(request.query, request.mentions)
    citations = [
        {
            "id": f"citation-{row['id']}",
            "entityId": row["id"],
            "entityType": row["type"],
            "label": row["title"],
            "description": row["subtitle"],
            "page": row.get("page"),
        }
        for row in rows[:8]
    ]

    return {
        "answerHint": "Найден контекст в Neo4j knowledge graph." if rows else "В Neo4j не найдено точных совпадений.",
        "citations": citations,
        "sourcesFound": len(citations),
        "experimentsFound": sum(1 for row in rows if row["type"] == "experiment"),
        "contextChunks": [
            f"{row['type']}: {row['title']} - {row['description']}"
            for row in rows[:12]
        ],
    }


@app.post("/internal/graphrag/extract")
def extract(request: ExtractRequest) -> dict[str, Any]:
    text = load_document_text(request)
    entities = extract_entities(request.documentId, text, request.title)
    relations = extract_relations(request.documentId, entities)

    return {
        "documentId": request.documentId,
        "entities": entities,
        "relations": relations,
        "warnings": [] if text else ["Не удалось извлечь текст из документа; создан минимальный draft."],
    }


@app.post("/internal/graphrag/publish")
def publish(request: PublishRequest) -> dict[str, Any]:
    extraction = request.extraction
    entities = extraction.get("entities", [])
    relations = extraction.get("relations", [])
    publish_to_neo4j(extraction.get("documentId"), entities, relations)

    return {
        "result": {
            "documentId": extraction.get("documentId"),
            "publishedEntityIds": [entity["id"] for entity in entities],
            "publishedRelationIds": [relation["id"] for relation in relations],
        }
    }


def load_document_text(request: ExtractRequest) -> str:
    if not request.storageKey:
        return request.title

    suffix = "." + request.type.lower().lstrip(".")
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        minio_client.fget_object(MINIO_BUCKET, request.storageKey, tmp.name)
        path = Path(tmp.name)

    try:
        if request.type == "pdf":
            return "\n".join(page.extract_text() or "" for page in PdfReader(str(path)).pages)
        if request.type == "docx":
            document = DocxDocument(str(path))
            return "\n".join(paragraph.text for paragraph in document.paragraphs)
        if request.type == "xlsx":
            workbook = load_workbook(str(path), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    rows.append(" ".join(str(cell) for cell in row if cell is not None))
            return "\n".join(rows)
        if request.type == "csv":
            with path.open("r", encoding="utf-8", errors="ignore") as file:
                return "\n".join(" ".join(row) for row in csv.reader(file))
        return path.read_text(encoding="utf-8", errors="ignore")
    finally:
        path.unlink(missing_ok=True)


def extract_entities(document_id: str, text: str, title: str) -> list[dict[str, Any]]:
    haystack = text or title
    material_match = re.search(r"(?:сплав|alloy)\s+([A-ZА-Я0-9-]+)", haystack, re.IGNORECASE)
    temperature_match = re.search(r"(\d{2,4})\s*(?:°C|C|С)", haystack)
    duration_match = re.search(r"(\d+(?:[,.]\d+)?)\s*(?:ч|час|hours?)", haystack, re.IGNORECASE)
    property_name = "Твердость" if re.search(r"HRC|HV|тверд", haystack, re.IGNORECASE) else "Предел прочности"
    material_name = f"Сплав {material_match.group(1)}" if material_match else "Материал из документа"

    entities = [
        {
            "id": f"{document_id}-material",
            "type": "material",
            "name": material_name,
            "attributes": extract_composition(haystack),
            "source": {"documentId": document_id, "page": 1},
        },
        {
            "id": f"{document_id}-experiment",
            "type": "experiment",
            "name": f"Эксперимент из {document_id}",
            "attributes": [
                {"name": "Температура", "value": int(temperature_match.group(1)), "unit": "°C"}
                if temperature_match
                else {"name": "Температура", "value": "Не указано", "unit": None},
                {"name": "Длительность", "value": duration_match.group(1).replace(",", "."), "unit": "ч"}
                if duration_match
                else {"name": "Длительность", "value": "Не указано", "unit": None},
            ],
            "source": {"documentId": document_id, "page": 1},
        },
        {
            "id": f"{document_id}-property",
            "type": "property",
            "name": property_name,
            "attributes": extract_measurements(haystack),
            "source": {"documentId": document_id, "page": 1},
        },
    ]

    if "метод" in haystack.lower() or "method" in haystack.lower():
        entities.append(
            {
                "id": f"{document_id}-unclassified-method",
                "type": "unclassified",
                "name": "Методика из документа",
                "attributes": [{"name": "Контекст", "value": "Методика требует ручной классификации", "unit": None}],
                "source": {"documentId": document_id, "page": 1},
            }
        )
    return entities


def extract_composition(text: str) -> list[dict[str, Any]]:
    return [
        {"name": element, "value": float(value.replace(",", ".")), "unit": "%"}
        for element, value in re.findall(r"\b([A-Z][a-z]?)\s*[-:=]?\s*(\d+(?:[,.]\d+)?)\s*%", text)
    ]


def extract_measurements(text: str) -> list[dict[str, Any]]:
    matches = re.findall(r"(\d+(?:[,.]\d+)?)\s*(HRC|HV|МПа|MPa)", text, re.IGNORECASE)
    return [
        {"name": f"Измерение {index + 1}", "value": float(value.replace(",", ".")), "unit": unit}
        for index, (value, unit) in enumerate(matches[:6])
    ]


def extract_relations(document_id: str, entities: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ids = {entity["type"]: entity["id"] for entity in entities}
    relations = [
        {
            "id": f"{document_id}-relation-material",
            "sourceId": ids["experiment"],
            "type": "USES_MATERIAL",
            "targetId": ids["material"],
            "source": {"documentId": document_id, "page": 1},
        },
        {
            "id": f"{document_id}-relation-property",
            "sourceId": ids["experiment"],
            "type": "MEASURES",
            "targetId": ids["property"],
            "source": {"documentId": document_id, "page": 1},
        },
    ]
    if "unclassified" in ids:
        relations.append(
            {
                "id": f"{document_id}-relation-unclassified",
                "sourceId": ids["experiment"],
                "type": "USES",
                "targetId": ids["unclassified"],
                "source": {"documentId": document_id, "page": 1},
            }
        )
    return relations


def publish_to_neo4j(document_id: str, entities: list[dict[str, Any]], relations: list[dict[str, Any]]) -> None:
    with driver.session() as session:
        session.run(
            """
            merge (d:Document {id: $id})
            set d.title = $id, d.type = 'document', d.subtitle = 'Исходный документ'
            """,
            id=document_id,
        )
        for entity in entities:
            session.run(
                """
                merge (e:KnowledgeEntity {id: $id})
                set e.type = $type,
                    e.title = $name,
                    e.subtitle = $type,
                    e.description = $description,
                    e.attributes = $attributes,
                    e.page = $page
                with e
                match (d:Document {id: $document_id})
                merge (e)-[:DESCRIBED_IN]->(d)
                """,
                id=entity["id"],
                type=entity["type"],
                name=entity["name"],
                description=f"Извлечено из документа {document_id}",
                attributes=str(entity.get("attributes", [])),
                page=(entity.get("source") or {}).get("page"),
                document_id=document_id,
            )
        for relation in relations:
            session.run(
                """
                match (source:KnowledgeEntity {id: $source_id})
                match (target:KnowledgeEntity {id: $target_id})
                merge (source)-[r:RELATED {id: $id}]->(target)
                set r.label = $label
                """,
                id=relation["id"],
                source_id=relation["sourceId"],
                target_id=relation["targetId"],
                label=relation["type"],
            )


def query_graph(query: str, mentions: list[EntityMention]) -> list[dict[str, Any]]:
    mention_ids = [mention.id for mention in mentions]
    normalized = query.lower()
    with driver.session() as session:
        result = session.run(
            """
            match (e:KnowledgeEntity)
            where e.id in $mention_ids
               or toLower(e.title) contains $query
               or toLower(e.description) contains $query
               or toLower(e.type) contains $query
            return e.id as id,
                   e.type as type,
                   e.title as title,
                   e.subtitle as subtitle,
                   e.description as description,
                   e.page as page
            limit 20
            """,
            mention_ids=mention_ids,
            query=normalized,
        )
        return [dict(record) for record in result]
